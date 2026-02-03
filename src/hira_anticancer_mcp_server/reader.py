"""
건강보험심사평가원(HIRA) 항암화학요법 파일 리더.

Excel(허가초과_항암요법)과 PDF(항암화학요법_공고전문)를 파싱하여
MCP TextContent / ImageContent 형태로 반환합니다.

전략:
  - Excel: openpyxl text extraction (머지셀 forward-fill, data_only=True)
  - PDF:   하이브리드 (텍스트 전용 → pdfplumber, 테이블 포함 → PyMuPDF ImageContent)
"""

from __future__ import annotations

import base64
import io
import logging
from pathlib import Path
from typing import Any

from mcp.types import ImageContent, TextContent

logger = logging.getLogger("hira-mcp-reader")

# ─────────────────────────────────────────────────────────────────────
# Excel 리더 (openpyxl)
# ─────────────────────────────────────────────────────────────────────

# 선호 시트 키워드 (sheet 미지정 시 이 키워드를 포함하는 시트를 우선 선택)
_PREFERRED_SHEET_KEYWORDS = ["인정", "용법용량"]

# 헤더 행 판별용 키워드 (이 중 2개 이상 포함 시 헤더로 간주)
_HEADER_KEYWORDS = ["요법코드", "암종", "항암화학요법", "투여대상", "투여단계",
                    "연번", "구분", "적응증", "약제", "성분명"]


def read_excel(
    filepath: Path,
    *,
    sheet: str | None = None,
    cancer_type: str | None = None,
    max_rows: int = 200,
) -> list[TextContent]:
    """
    Excel 파일을 읽어 Markdown 테이블로 변환합니다.

    Args:
        filepath: .xlsx 파일 경로
        sheet: 시트 이름 (None이면 자동 선택)
        cancer_type: 암종 필터 (예: "난소암", "자궁경부암")
        max_rows: 최대 반환 행 수 (토큰 제한 방지)

    Returns:
        list[TextContent] — Markdown 테이블 + 요약 정보
    """
    import openpyxl

    wb = openpyxl.load_workbook(str(filepath), data_only=True, read_only=False)

    if sheet:
        if sheet not in wb.sheetnames:
            return [TextContent(
                type="text",
                text=f"⚠️ 시트 '{sheet}'를 찾을 수 없습니다.\n"
                     f"사용 가능한 시트: {', '.join(wb.sheetnames)}"
            )]
        ws = wb[sheet]
    else:
        ws = _select_preferred_sheet(wb)

    # ── 머지셀 forward-fill 맵 구축 ─────────────────────────────
    merge_map: dict[tuple[int, int], Any] = {}
    for merged_range in ws.merged_cells.ranges:
        top_left_value = ws.cell(
            row=merged_range.min_row,
            column=merged_range.min_col,
        ).value
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merge_map[(row, col)] = top_left_value

    # ── 데이터 추출 ─────────────────────────────────────────────
    all_rows: list[list[str]] = []
    for row_idx, row in enumerate(ws.iter_rows(values_only=False), start=1):
        cells: list[str] = []
        for cell in row:
            coord = (cell.row, cell.column)
            if coord in merge_map:
                val = merge_map[coord]
            else:
                val = cell.value
            cells.append(str(val).strip() if val is not None else "")
        all_rows.append(cells)

    sheet_title = ws.title
    sheet_names = wb.sheetnames
    wb.close()

    if not all_rows:
        return [TextContent(type="text", text="⚠️ 시트에 데이터가 없습니다.")]

    # ── 헤더 감지 (키워드 기반) ──────────────────────────────────
    header_idx = _find_header_row(all_rows)

    headers = all_rows[header_idx]
    data_rows = all_rows[header_idx + 1:]

    # 빈 행 제거 (모든 셀이 비어있는 행)
    data_rows = [row for row in data_rows if any(c for c in row)]

    # ── 암종 필터 적용 ──────────────────────────────────────────
    if cancer_type:
        cancer_col_idx = _find_cancer_column(headers)
        if cancer_col_idx is not None:
            data_rows = [
                row for row in data_rows
                if cancer_type in row[cancer_col_idx]
            ]

    total_count = len(data_rows)
    truncated = total_count > max_rows
    data_rows = data_rows[:max_rows]

    # ── Markdown 테이블 생성 ────────────────────────────────────
    md_lines = _to_markdown_table(headers, data_rows)

    # ── 요약 정보 ──────────────────────────────────────────────
    summary_parts = [
        f"📊 시트: {sheet_title}",
        f"📏 전체 행: {total_count}행",
    ]
    if cancer_type:
        summary_parts.append(f"🔍 필터: '{cancer_type}'")
    if truncated:
        summary_parts.append(f"⚠️ {max_rows}행까지만 표시 (전체 {total_count}행)")

    summary = " | ".join(summary_parts)
    sheets_info = f"사용 가능한 시트: {', '.join(sheet_names)}"

    # ── 검토중/불승인 시트 경고 메시지 ─────────────────────────
    warning = ""
    if "검토중" in sheet_title:
        warning = (
            "\n\n🚨 **주의**: 이 시트의 항암요법은 현재 **검토 중**입니다. "
            "아직 건강보험 급여로 승인되지 않았으며, 향후 변경될 수 있습니다. "
            "승인된 요법은 '인정되고 있는 허가초과 항암요법(용법용량포함)' 시트를 확인하세요."
        )
    elif "불승인" in sheet_title:
        warning = (
            "\n\n🚨 **주의**: 이 시트의 항암요법은 **승인 거부(불승인)**되었습니다. "
            "건강보험 급여로 인정되지 않습니다. "
            "승인된 요법은 '인정되고 있는 허가초과 항암요법(용법용량포함)' 시트를 확인하세요."
        )

    return [TextContent(type="text", text=f"{summary}\n{sheets_info}{warning}\n\n{md_lines}")]


def _select_preferred_sheet(wb):
    """선호 시트를 자동 선택합니다. 키워드 매칭 → 활성 시트 순."""
    for name in wb.sheetnames:
        if all(kw in name for kw in _PREFERRED_SHEET_KEYWORDS):
            logger.info(f"선호 시트 자동 선택: {name}")
            return wb[name]
    return wb.active


def _find_header_row(all_rows: list[list[str]]) -> int:
    """헤더 키워드가 포함된 행을 찾습니다. 없으면 첫 비어있지 않은 행."""
    for i, row in enumerate(all_rows):
        row_text = " ".join(row).lower()
        matches = sum(1 for kw in _HEADER_KEYWORDS if kw in row_text)
        if matches >= 2:
            return i

    # fallback: 첫 번째 비어있지 않은 행
    for i, row in enumerate(all_rows):
        if any(c for c in row):
            return i
    return 0


def _find_cancer_column(headers: list[str]) -> int | None:
    """헤더에서 암종 관련 컬럼 인덱스를 찾습니다."""
    cancer_keywords = ["암종", "cancer", "질환", "적응증", "진단", "암 종"]
    for idx, h in enumerate(headers):
        h_lower = h.lower()
        if any(kw in h_lower for kw in cancer_keywords):
            return idx
    # fallback: "투여대상" 컬럼 (암종명이 투여대상에 포함되는 경우도 있음)
    for idx, h in enumerate(headers):
        if "투여대상" in h:
            return idx
    return None


def _to_markdown_table(headers: list[str], rows: list[list[str]]) -> str:
    """헤더와 데이터 행을 Markdown 테이블 문자열로 변환합니다."""
    if not headers:
        return "(빈 테이블)"

    # 열 수 통일
    n_cols = len(headers)

    # 긴 셀 내용 축약 (300자 초과 시)
    def _trunc(s: str, limit: int = 300) -> str:
        return s[:limit] + "…" if len(s) > limit else s

    header_line = "| " + " | ".join(_trunc(h) for h in headers) + " |"
    sep_line = "| " + " | ".join("---" for _ in headers) + " |"

    data_lines = []
    for row in rows:
        # 열 수가 헤더보다 적으면 빈 문자열로 채움
        padded = row[:n_cols] + [""] * max(0, n_cols - len(row))
        line = "| " + " | ".join(_trunc(c) for c in padded) + " |"
        data_lines.append(line)

    return "\n".join([header_line, sep_line] + data_lines)


# ─────────────────────────────────────────────────────────────────────
# PDF 리더 (pdfplumber + PyMuPDF 하이브리드)
# ─────────────────────────────────────────────────────────────────────

# 페이지 타입 감지 상수
_TABLE_THRESHOLD = 1  # pdfplumber가 N개 이상 테이블 감지 시 → 이미지 렌더링
_MAX_PAGES_PER_CALL = 50  # 한 번에 처리할 최대 페이지 (토큰 제한 방지)
_IMAGE_DPI = 150  # ImageContent 해상도
_MAX_IMAGE_PAGES = 5  # 이미지 렌더링 최대 페이지 (1MB 제한 방지)

# 암종 영한 매핑 (검색용)
_CANCER_ALIASES: dict[str, list[str]] = {
    "소세포폐암": ["small cell lung", "sclc"],
    "비소세포폐암": ["non-small cell lung", "nsclc"],
    "위암": ["gastric", "stomach"],
    "식도암": ["esophageal", "esophagus"],
    "갑상선암": ["thyroid"],
    "췌장암": ["pancreatic", "pancreas"],
    "간암": ["hepatocellular", "liver", "hcc"],
    "담도암": ["biliary", "cholangiocarcinoma"],
    "직결장암": ["colorectal", "colon", "rectal", "crc"],
    "유방암": ["breast"],
    "난소암": ["ovarian", "ovary"],
    "난관암": ["fallopian"],
    "자궁경부암": ["cervical", "cervix"],
    "자궁암": ["uterine", "endometrial"],
    "자궁내막암": ["endometrial", "endometrium"],
    "신장암": ["renal", "kidney", "rcc"],
    "요로상피암": ["urothelial", "bladder"],
    "전립선암": ["prostate"],
    "두경부암": ["head and neck", "head & neck"],
    "신경내분비암": ["neuroendocrine", "net"],
    "메르켈세포암": ["merkel"],
    "피부암": ["skin", "bcc", "scc"],
    "골암": ["bone", "osteosarcoma"],
    "중추신경계암": ["cns", "brain", "glioma", "glioblastoma"],
    "악성흑색종": ["melanoma"],
    "연조직육종": ["soft tissue sarcoma"],
    "횡문근육종": ["rhabdomyosarcoma"],
    "생식세포종양": ["germ cell"],
    "신경모세포종": ["neuroblastoma"],
    "윌름즈종양": ["wilms"],
    "망막모세포종": ["retinoblastoma"],
    "비호지킨림프종": ["non-hodgkin", "nhl", "lymphoma"],
    "호지킨림프종": ["hodgkin"],
    "다발골수종": ["multiple myeloma", "myeloma"],
    "급성골수성백혈병": ["aml", "acute myeloid"],
    "급성전골수구성백혈병": ["apl", "promyelocytic"],
    "만성골수성백혈병": ["cml", "chronic myeloid"],
    "급성림프모구백혈병": ["all", "acute lymphoblastic"],
    "만성림프구성백혈병": ["cll", "chronic lymphocytic"],
    "골수형성이상증후군": ["mds", "myelodysplastic"],
}

# PDF 섹션별 키워드 매핑 (항암화학요법 공고전문 구조)
PDF_SECTIONS: dict[str, list[str]] = {
    "일반원칙": ["일반원칙"],
    "암종별항암요법": ["주요 암종별 항암요법"],
    "항암면역요법제": ["항암면역요법제"],
    "항구토제": ["항구토제"],
    "별표": ["별표", "[별표"],
    "부록": ["부록", "부표"],
}


def read_pdf(
    filepath: Path,
    *,
    pages: str | None = None,
    section: str | None = None,
    cancer_type: str | None = None,
    search: str | None = None,
    text_only: bool = False,
) -> list[TextContent | ImageContent]:
    """
    PDF를 하이브리드 방식으로 읽습니다.

    Args:
        filepath: .pdf 파일 경로
        pages: 페이지 범위 (예: "1-10", "5", "1,3,7-10"). None이면 처음 50p.
        section: 섹션 필터 (예: "일반원칙", "별표").
        cancer_type: 암종명 (예: "난소암", "ovarian"). TOC에서 페이지 범위 자동 탐색.
        search: 키워드 검색 (예: 약제명, 암종명). 매칭 페이지와 주변 텍스트 반환.
        text_only: True이면 이미지 없이 텍스트만 반환 (1MB 제한 회피).

    Returns:
        list[TextContent | ImageContent] 혼합 리스트
    """
    import fitz  # PyMuPDF
    import pdfplumber

    doc = fitz.open(str(filepath))
    total_pages = len(doc)

    # ── 키워드 검색 모드 ────────────────────────────────────────
    if search:
        doc.close()
        return _search_pdf(filepath, search, total_pages)

    # ── 페이지 범위 결정 ────────────────────────────────────────
    range_label = None  # 사용자에게 보여줄 범위 설명

    if cancer_type:
        toc, toc_page_idx = _parse_toc(filepath)
        page_indices, matched_name = _find_cancer_pages(toc, cancer_type, total_pages, filepath, toc_page_idx)
        if not page_indices:
            doc.close()
            available = ", ".join(e["name"] for e in toc) if toc else "(TOC 파싱 실패)"
            return [TextContent(
                type="text",
                text=f"⚠️ 암종 '{cancer_type}'을 목차에서 찾을 수 없습니다.\n"
                     f"사용 가능한 암종: {available}"
            )]
        range_label = f"🔍 암종: '{matched_name}'"
    elif section:
        toc, toc_page_idx = _parse_toc(filepath)
        page_indices = _find_section_pages_from_toc(toc, section, filepath, total_pages, toc_page_idx)
        if not page_indices:
            doc.close()
            return [TextContent(
                type="text",
                text=f"⚠️ 섹션 '{section}'을 찾을 수 없습니다.\n"
                     f"사용 가능한 섹션: {', '.join(PDF_SECTIONS.keys())}\n"
                     f"총 {total_pages}페이지"
            )]
        range_label = f"🔍 섹션: '{section}'"
    elif pages:
        page_indices = _parse_page_range(pages, total_pages)
    else:
        # 기본: TOC 페이지를 보여줌 (사용자가 탐색할 수 있도록)
        toc, _toc_idx = _parse_toc(filepath)
        if toc:
            doc.close()
            return _format_toc_response(filepath, toc, total_pages)
        page_indices = list(range(min(total_pages, _MAX_PAGES_PER_CALL)))

    # 50페이지 제한 적용
    truncated = len(page_indices) > _MAX_PAGES_PER_CALL
    page_indices = page_indices[:_MAX_PAGES_PER_CALL]

    # ── 이미지 페이지 수 제한 (1MB 방지) ────────────────────────
    # text_only가 아닌 경우에도 이미지 페이지 수를 제한
    image_page_count = 0

    # ── 페이지별 타입 감지 + 파싱 ─────────────────────────────
    results: list[TextContent | ImageContent] = []

    # 시작 메타 정보
    meta = (
        f"📄 PDF: {filepath.name} ({total_pages}p)\n"
        f"📖 표시 범위: {_format_page_range(page_indices)} "
        f"({len(page_indices)}p)"
    )
    if truncated:
        meta += f"\n⚠️ {_MAX_PAGES_PER_CALL}p 제한 적용됨"
    if range_label:
        meta += f"\n{range_label}"
    if text_only:
        meta += "\n📝 텍스트 전용 모드"
    results.append(TextContent(type="text", text=meta))

    # pdfplumber로 테이블 감지
    pdf_plumber = pdfplumber.open(str(filepath))

    text_buffer: list[str] = []  # 연속 텍스트 페이지 버퍼

    for page_idx in page_indices:
        page_num = page_idx + 1  # 1-indexed

        try:
            plumber_page = pdf_plumber.pages[page_idx]
        except IndexError:
            continue

        # text_only 모드이면 항상 텍스트 추출
        if text_only:
            text = _extract_text_safe(plumber_page, page_num)
            text_buffer.append(text)
            continue

        # pdfplumber로 테이블 감지
        try:
            tables = plumber_page.find_tables()
            has_tables = len(tables) >= _TABLE_THRESHOLD
        except Exception:
            has_tables = False

        if has_tables and image_page_count < _MAX_IMAGE_PAGES:
            # 텍스트 버퍼가 있으면 먼저 flush
            if text_buffer:
                results.append(TextContent(
                    type="text", text="\n\n".join(text_buffer)
                ))
                text_buffer.clear()

            # 테이블 페이지 → 이미지 렌더링 (PyMuPDF)
            try:
                fitz_page = doc[page_idx]
                mat = fitz.Matrix(_IMAGE_DPI / 72, _IMAGE_DPI / 72)
                pix = fitz_page.get_pixmap(matrix=mat)
                png_bytes = pix.tobytes("png")

                b64_data = base64.b64encode(png_bytes).decode("ascii")

                results.append(TextContent(
                    type="text",
                    text=f"--- 📊 p.{page_num} (테이블 포함 → 이미지) ---"
                ))
                results.append(ImageContent(
                    type="image",
                    data=b64_data,
                    mimeType="image/png",
                ))
                image_page_count += 1
            except Exception as e:
                logger.warning(f"페이지 {page_num} 이미지 렌더링 실패: {e}")
                text = _extract_text_safe(plumber_page, page_num)
                text_buffer.append(text)
        else:
            # 텍스트 전용 페이지 또는 이미지 제한 초과
            if has_tables and image_page_count >= _MAX_IMAGE_PAGES:
                text = _extract_text_safe(plumber_page, page_num)
                text_buffer.append(
                    f"--- p.{page_num} (테이블 포함, 이미지 제한 초과 → 텍스트) ---\n"
                    + text.split("\n", 1)[-1] if "\n" in text else text
                )
            else:
                text = _extract_text_safe(plumber_page, page_num)
                text_buffer.append(text)

    # 남은 텍스트 버퍼 flush
    if text_buffer:
        results.append(TextContent(
            type="text", text="\n\n".join(text_buffer)
        ))

    if image_page_count >= _MAX_IMAGE_PAGES:
        results.append(TextContent(
            type="text",
            text=f"\n⚠️ 이미지 렌더링 {_MAX_IMAGE_PAGES}p 제한 도달. "
                 f"나머지 테이블 페이지는 텍스트로 반환됨. "
                 f"text_only=true로 전체 텍스트 조회 가능."
        ))

    pdf_plumber.close()
    doc.close()

    return results


def _extract_text_safe(plumber_page, page_num: int) -> str:
    """pdfplumber 페이지에서 안전하게 텍스트를 추출합니다."""
    try:
        text = plumber_page.extract_text() or ""
        if text.strip():
            return f"--- p.{page_num} ---\n{text.strip()}"
        else:
            return f"--- p.{page_num} (빈 페이지) ---"
    except Exception as e:
        return f"--- p.{page_num} (추출 실패: {e}) ---"


# ─────────────────────────────────────────────────────────────────────
# PDF TOC 파싱 (목차에서 암종→페이지 매핑 추출)
# ─────────────────────────────────────────────────────────────────────
import re

# 항목 시작 패턴: "숫자. " 또는 "숫자-숫자. "
_TOC_ENTRY_START = re.compile(r"(\d+(?:-\d+)?)\.\s")

# 섹션 레벨 패턴: "□ 섹션명···숫자"
_TOC_SECTION_PATTERN = re.compile(r"□\s*(.+?)·+\s*(\d+)")


def _parse_toc_entries_from_line(line: str) -> list[dict]:
    """한 줄에서 TOC 항목들을 추출합니다 (두 컬럼 대응)."""
    entries = []
    # 항목 시작 위치 찾기
    starts = list(_TOC_ENTRY_START.finditer(line))
    for i, match in enumerate(starts):
        num = match.group(1)
        text_start = match.end()
        # 다음 항목 시작 또는 줄 끝까지가 이 항목의 텍스트
        text_end = starts[i + 1].start() if i + 1 < len(starts) else len(line)
        segment = line[text_start:text_end].strip()

        # segment에서 이름과 페이지 번호 분리
        # 패턴: "이름·····숫자" 또는 "이름 숫자" (마지막 숫자가 페이지)
        # 먼저 dot 구분 시도 (첫 번째 ·+숫자 매칭 — 비탐욕적)
        dot_match = re.match(r"(.+?)·+(\d+)", segment)
        if dot_match:
            name = dot_match.group(1).strip()
            page = int(dot_match.group(2))
        else:
            # dot 없는 경우: 마지막 숫자를 페이지로 추출
            num_match = re.search(r"\s(\d+)\s*$", segment)
            if num_match:
                name = segment[:num_match.start()].strip()
                page = int(num_match.group(1))
            else:
                continue  # 파싱 실패 → 건너뜀

        name = re.sub(r"\s+", " ", name)
        if name and page > 0:
            entries.append({"num": num, "name": name, "page": page})

    return entries


def _parse_toc(filepath: Path) -> tuple[list[dict], int]:
    """
    PDF 목차 페이지를 파싱하여 암종별 페이지 매핑을 추출합니다.

    Returns:
        (entries, toc_page_idx) where entries is
        [{"num": "1", "name": "소세포폐암", "page": 16}, ...]
        페이지 번호 순으로 정렬됨. toc_page_idx는 목차 페이지의 실제 PDF 인덱스.
    """
    import pdfplumber

    pdf = pdfplumber.open(str(filepath))
    toc_entries: list[dict] = []
    section_entries: list[dict] = []

    # 목차 페이지 탐색 — 가장 많은 항목이 있는 페이지를 선택
    best_page_idx = -1
    best_count = 0
    for i in range(25, min(50, len(pdf.pages))):
        text = pdf.pages[i].extract_text() or ""
        if "일반원칙" in text and "암종별" in text:
            count = len(list(_TOC_ENTRY_START.finditer(text)))
            if count > best_count:
                best_count = count
                best_page_idx = i

    if best_page_idx >= 0:
        text = pdf.pages[best_page_idx].extract_text() or ""

        # 섹션 레벨 항목 추출
        for match in _TOC_SECTION_PATTERN.finditer(text):
            name, page = match.group(1).strip(), int(match.group(2))
            section_entries.append({"name": name, "page": page})

        # 줄 단위로 암종 항목 추출
        for line in text.split("\n"):
            line = line.strip()
            if not line or line.startswith("□") or line.startswith("암환자"):
                continue
            entries = _parse_toc_entries_from_line(line)
            toc_entries.extend(entries)

    pdf.close()

    # 페이지 번호 순 정렬 (두 컬럼이 섞여있으므로)
    toc_entries.sort(key=lambda e: e["page"])

    # 중복 제거 (같은 페이지)
    seen = set()
    unique = []
    for entry in toc_entries:
        if entry["page"] not in seen:
            seen.add(entry["page"])
            unique.append(entry)
    toc_entries = unique

    # 각 항목의 end_page 계산 (다음 항목의 시작 - 1)
    for i, entry in enumerate(toc_entries):
        if i + 1 < len(toc_entries):
            entry["end_page"] = toc_entries[i + 1]["page"] - 1
        else:
            # 마지막 암종 항목: "항암면역요법제" 섹션 시작 전까지
            next_section_page = None
            for sec in section_entries:
                if sec["page"] > entry["page"]:
                    next_section_page = sec["page"]
                    break
            entry["end_page"] = (next_section_page - 1) if next_section_page else entry["page"] + 10

    logger.info(f"TOC 파싱 완료: {len(toc_entries)}개 항목, TOC page idx={best_page_idx}")
    return toc_entries, best_page_idx


def _find_cancer_pages(
    toc: list[dict], cancer_type: str, total_pages: int,
    filepath: Path | None = None,
    toc_page_idx: int = -1,
) -> tuple[list[int], str]:
    """
    TOC에서 암종명으로 페이지 범위를 찾습니다. 퍼지 매칭 지원.

    Returns:
        (page_indices, matched_name) — 못 찾으면 ([], "")
    """
    query = cancer_type.lower().strip()

    def _resolve(entry: dict) -> tuple[list[int], str]:
        start, end = _toc_page_to_indices(entry, toc, total_pages, filepath, toc_page_idx)
        return list(range(start, end + 1)), entry["name"]

    # 1단계: 정확한 한글 이름 매칭
    for entry in toc:
        if query in entry["name"]:
            return _resolve(entry)

    # 2단계: 영문 별칭 매칭
    for korean_name, aliases in _CANCER_ALIASES.items():
        if query in korean_name or any(alias in query for alias in aliases):
            for entry in toc:
                if korean_name in entry["name"]:
                    return _resolve(entry)

    # 3단계: 부분 매칭 (가장 유사한 항목)
    for entry in toc:
        entry_lower = entry["name"].lower()
        if any(c in entry_lower for c in query if len(c) > 1):
            return _resolve(entry)

    return [], ""


_toc_offset_cache: dict[str, int] = {}


def _calc_toc_offset(
    filepath: Path, toc: list[dict], toc_page_idx: int = -1
) -> int:
    """
    TOC 페이지 번호와 실제 PDF 페이지의 오프셋을 계산합니다.

    방법: TOC 직후 첫 콘텐츠 페이지의 하단 인쇄 페이지 번호를 읽어서
    offset = pdf_idx - printed_number + 1 로 계산.
    """
    cache_key = str(filepath)
    if cache_key in _toc_offset_cache:
        return _toc_offset_cache[cache_key]

    import pdfplumber

    pdf = pdfplumber.open(str(filepath))

    # 방법 1: TOC 직후 페이지의 footer 번호로 오프셋 계산
    if toc_page_idx >= 0:
        for scan_idx in range(toc_page_idx + 1, min(toc_page_idx + 5, len(pdf.pages))):
            text = pdf.pages[scan_idx].extract_text() or ""
            lines = [ln.strip() for ln in text.strip().split("\n") if ln.strip()]
            if not lines:
                continue
            # footer: 마지막 줄이 숫자만 있는 경우
            last_line = lines[-1]
            footer_match = re.match(r"^(\d+)$", last_line)
            if footer_match:
                footer_num = int(footer_match.group(1))
                offset = scan_idx - footer_num + 1
                _toc_offset_cache[cache_key] = offset
                pdf.close()
                logger.info(
                    f"TOC 오프셋 계산 (footer): {offset} "
                    f"(PDF idx={scan_idx}, footer={footer_num})"
                )
                return offset

    # 방법 2 (fallback): "일반원칙" 텍스트 위치 + TOC/section 항목 대조
    for i in range(30, min(50, len(pdf.pages))):
        text = (pdf.pages[i].extract_text() or "")[:500]
        if "일반원칙" in text:
            # footer 번호 확인
            lines = [ln.strip() for ln in text.strip().split("\n") if ln.strip()]
            footer_match = re.match(r"^(\d+)$", lines[-1]) if lines else None
            if footer_match:
                footer_num = int(footer_match.group(1))
                offset = i - footer_num + 1
                _toc_offset_cache[cache_key] = offset
                pdf.close()
                logger.info(f"TOC 오프셋 계산 (일반원칙 fallback): {offset}")
                return offset

    pdf.close()

    # 최후 fallback
    _toc_offset_cache[cache_key] = 33
    logger.warning("TOC 오프셋 계산 실패, 기본값 33 사용")
    return 33


def _toc_page_to_indices(
    entry: dict, toc: list[dict], total_pages: int,
    filepath: Path | None = None,
    toc_page_idx: int = -1,
) -> tuple[int, int]:
    """
    TOC 페이지 번호(PDF 내부 번호)를 0-indexed 페이지 인덱스로 변환합니다.
    """
    offset = _calc_toc_offset(filepath, toc, toc_page_idx) if filepath else 33

    start_idx = entry["page"] + offset - 1  # 0-indexed
    end_idx = entry["end_page"] + offset - 1

    # ±2 페이지 퍼지 검증: 암종명이 실제 해당 페이지에 있는지 확인
    if filepath and entry.get("name"):
        start_idx = _verify_page_with_fuzzy(
            filepath, start_idx, entry["name"], total_pages
        )
        # end도 조정 (start와의 차이 유지)
        page_span = entry["end_page"] - entry["page"]
        end_idx = start_idx + page_span

    # 범위 검증
    start_idx = max(0, min(start_idx, total_pages - 1))
    end_idx = max(start_idx, min(end_idx, total_pages - 1))

    return start_idx, end_idx


def _verify_page_with_fuzzy(
    filepath: Path, expected_idx: int, cancer_name: str, total_pages: int,
    search_range: int = 2,
) -> int:
    """
    예상 페이지 ±search_range 범위에서 암종명을 검색하여 실제 시작 페이지를 반환합니다.
    찾지 못하면 원래 expected_idx를 반환합니다.
    """
    import pdfplumber

    # 짧은 이름 추출 (예: "난소암/난관암/일차복막암" → ["난소암", "난관암"])
    name_parts = [p.strip() for p in cancer_name.replace("/", "|").split("|") if len(p.strip()) >= 2]
    if not name_parts:
        return expected_idx

    pdf = pdfplumber.open(str(filepath))
    try:
        # 예상 페이지 먼저 확인
        if 0 <= expected_idx < total_pages:
            text = (pdf.pages[expected_idx].extract_text() or "")[:500]
            if any(part in text for part in name_parts):
                return expected_idx

        # ±search_range 탐색
        for delta in range(1, search_range + 1):
            for candidate in [expected_idx + delta, expected_idx - delta]:
                if 0 <= candidate < total_pages:
                    text = (pdf.pages[candidate].extract_text() or "")[:500]
                    if any(part in text for part in name_parts):
                        logger.info(
                            f"퍼지 검증: '{cancer_name}' 페이지 조정 "
                            f"{expected_idx} → {candidate}"
                        )
                        return candidate
    finally:
        pdf.close()

    return expected_idx


def _find_section_pages_from_toc(
    toc: list[dict], section: str, filepath: Path, total_pages: int,
    toc_page_idx: int = -1,
) -> list[int]:
    """TOC 기반으로 섹션 페이지 범위를 찾습니다. 실패 시 텍스트 스캔 폴백."""
    keywords = PDF_SECTIONS.get(section, [section])

    # TOC에서 검색
    for entry in toc:
        if any(kw in entry["name"] for kw in keywords):
            start, end = _toc_page_to_indices(entry, toc, total_pages, filepath, toc_page_idx)
            return list(range(start, end + 1))

    # 폴백: 텍스트 스캔
    return _find_section_pages_by_scan(filepath, section, total_pages)


def _find_section_pages_by_scan(
    filepath: Path, section: str, total_pages: int
) -> list[int]:
    """PDF 전체를 스캔하여 섹션 페이지를 찾습니다 (폴백)."""
    import pdfplumber

    keywords = PDF_SECTIONS.get(section, [section])
    pdf = pdfplumber.open(str(filepath))
    start_page = None

    for i, page in enumerate(pdf.pages):
        text = (page.extract_text() or "").strip()
        if not text:
            continue
        header = text[:500]
        if any(kw in header for kw in keywords):
            start_page = i
            break

    if start_page is None:
        pdf.close()
        return []

    # 다음 섹션 시작점 탐색
    end_page = min(start_page + 50, total_pages - 1)
    other_keywords = []
    for sec_name, sec_kws in PDF_SECTIONS.items():
        if sec_name != section:
            other_keywords.extend(sec_kws)

    for i in range(start_page + 1, min(start_page + 100, total_pages)):
        text = (pdf.pages[i].extract_text() or "").strip()
        if any(kw in text[:500] for kw in other_keywords):
            end_page = i - 1
            break

    pdf.close()
    return list(range(start_page, end_page + 1))


# ─────────────────────────────────────────────────────────────────────
# PDF 키워드 검색
# ─────────────────────────────────────────────────────────────────────
_SEARCH_MAX_RESULTS = 20
_SEARCH_CONTEXT_CHARS = 200


def _search_pdf(
    filepath: Path, keyword: str, total_pages: int
) -> list[TextContent]:
    """PDF 전체에서 키워드를 검색하여 매칭 페이지와 주변 텍스트를 반환합니다."""
    import pdfplumber

    pdf = pdfplumber.open(str(filepath))
    matches: list[dict] = []
    keyword_lower = keyword.lower()

    for i, page in enumerate(pdf.pages):
        if len(matches) >= _SEARCH_MAX_RESULTS:
            break
        text = page.extract_text() or ""
        if keyword_lower in text.lower():
            # 매칭 위치의 주변 텍스트 추출
            idx = text.lower().index(keyword_lower)
            start = max(0, idx - _SEARCH_CONTEXT_CHARS)
            end = min(len(text), idx + len(keyword) + _SEARCH_CONTEXT_CHARS)
            context = text[start:end].strip()
            if start > 0:
                context = "…" + context
            if end < len(text):
                context = context + "…"
            matches.append({"page": i + 1, "context": context})

    pdf.close()

    if not matches:
        return [TextContent(
            type="text",
            text=f"🔍 '{keyword}' 검색 결과: 0건 (전체 {total_pages}p 검색)\n"
                 "다른 키워드나 영문/한글 변형을 시도해보세요."
        )]

    lines = [
        f"🔍 '{keyword}' 검색 결과: {len(matches)}건 "
        f"(전체 {total_pages}p 검색)",
        "─" * 40,
    ]
    for m in matches:
        lines.append(f"\n📍 p.{m['page']}:")
        lines.append(m["context"])

    lines.append("\n─" * 40)
    lines.append(
        "💡 특정 페이지를 자세히 보려면 pages 파라미터를 사용하세요. "
        "예: pages='" + ",".join(str(m["page"]) for m in matches[:5]) + "'"
    )

    return [TextContent(type="text", text="\n".join(lines))]


def _format_toc_response(
    filepath: Path, toc: list[dict], total_pages: int
) -> list[TextContent]:
    """TOC를 보기 좋게 포맷하여 반환합니다."""
    lines = [
        f"📄 PDF: {filepath.name} ({total_pages}p)",
        "",
        "📋 목차 (cancer_type 파라미터로 암종별 조회 가능):",
        "─" * 50,
    ]
    for entry in toc:
        lines.append(f"  {entry['num']:>5}. {entry['name']:<20} → p.{entry['page']}")

    lines.append("─" * 50)
    lines.append("")
    lines.append("💡 사용법:")
    lines.append("  • cancer_type='난소암' → 해당 암종 페이지 자동 조회")
    lines.append("  • search='trastuzumab' → 전체 PDF에서 키워드 검색")
    lines.append("  • pages='64-68' → 특정 페이지 범위 직접 조회")
    lines.append("  • text_only=true → 이미지 없이 텍스트만 (넓은 범위 조회)")

    return [TextContent(type="text", text="\n".join(lines))]


def _parse_page_range(pages_str: str, total: int) -> list[int]:
    """
    페이지 범위 문자열을 0-indexed 리스트로 변환합니다.

    지원 형식:
      - "5"       → [4]
      - "1-10"    → [0,1,...,9]
      - "1,3,7-10" → [0,2,6,7,8,9]
    """
    result: list[int] = []
    for part in pages_str.split(","):
        part = part.strip()
        if "-" in part:
            start_s, end_s = part.split("-", 1)
            start = max(int(start_s.strip()) - 1, 0)
            end = min(int(end_s.strip()) - 1, total - 1)
            result.extend(range(start, end + 1))
        else:
            idx = int(part) - 1
            if 0 <= idx < total:
                result.append(idx)

    return sorted(set(result))


def _format_page_range(indices: list[int]) -> str:
    """0-indexed 리스트를 사람이 읽기 좋은 페이지 범위로 변환합니다."""
    if not indices:
        return "(없음)"

    # 연속 구간 탐지
    ranges: list[str] = []
    start = indices[0]
    prev = indices[0]

    for i in indices[1:]:
        if i == prev + 1:
            prev = i
        else:
            if start == prev:
                ranges.append(f"p.{start + 1}")
            else:
                ranges.append(f"p.{start + 1}-{prev + 1}")
            start = i
            prev = i

    if start == prev:
        ranges.append(f"p.{start + 1}")
    else:
        ranges.append(f"p.{start + 1}-{prev + 1}")

    return ", ".join(ranges)
